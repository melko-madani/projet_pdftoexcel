"""Pipeline complet de structuration de dossiers de demandes fiscales.

Orchestre : dézippage → regroupement par préfixe → extraction métadonnées
→ extraction tableaux → génération Excel → construction ZIP de sortie.
"""

import io
import logging
import re
import tempfile
import zipfile
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import Workbook

from .excel_writer import (
    CELL_ALIGNMENT,
    EURO_FORMAT,
    HEADER_ALIGNMENT,
    HEADER_FILL,
    HEADER_FONT,
    META_VAL_FONT,
    THIN_BORDER,
    _sanitize_sheet_name,
    write_dataset_to_sheet,
    write_metadata_sheet,
)
from openpyxl.utils import get_column_letter
from .metadata import (
    DossierMetadata,
    RawMetadata,
    ComputedMetadata,
    build_raw_metadata,
    compute_metadata,
    computed_metadata_to_rows,
    detect_pdf_type,
    extract_metadata,
    metadata_to_rows,
)
from scripts.table_data_extractor import extract_from_datasets as extract_table_data
from .parser import Dataset, process_tables
from .scanner import scan_pdf

logger = logging.getLogger(__name__)

MIN_COLS_DEFAULT = 8


@dataclass
class DemandeDossier:
    """Un groupe de fichiers pour une même demande."""
    prefix: str
    courrier_pdf: bytes | None = None
    courrier_filename: str | None = None
    ar_pdf: bytes | None = None
    ar_filename: str | None = None
    depot_pdf: bytes | None = None
    depot_filename: str | None = None


@dataclass
class DemandeResult:
    """Résultat du traitement d'une demande."""
    prefix: str
    numero_demande: str = ""
    metadata: DossierMetadata = field(default_factory=DossierMetadata)
    raw_metadata: RawMetadata = field(default_factory=RawMetadata)
    computed_metadata: ComputedMetadata = field(default_factory=ComputedMetadata)
    datasets: list[Dataset] = field(default_factory=list)
    annexe_excel: bytes | None = None
    metadata_excel: bytes | None = None
    source_pdfs: dict[str, bytes] = field(default_factory=dict)
    source_filenames: dict[str, str] = field(default_factory=dict)
    table_count: int = 0
    row_count: int = 0
    error: str | None = None


def _extract_prefix(filename: str) -> str:
    """Extrait le préfixe numérique d'un nom de fichier."""
    match = re.match(r"^(\d+)", filename)
    return match.group(1) if match else ""


def _classify_file(filename: str, parent_dir: str) -> str:
    """Classifie un fichier en courrier/ar/depot basé sur le dossier parent et le nom."""
    parent_lower = parent_dir.lower()
    name_lower = filename.lower()

    # Fichiers dans mails/ → toujours courrier
    if "mail" in parent_lower:
        return "courrier"

    # Fichiers dans proof/ → AR ou dépôt
    if "proof" in parent_lower:
        if "ar_n" in name_lower or "ar " in name_lower or name_lower.startswith("ar_"):
            return "ar"
        if "preuve" in name_lower or "dépôt" in name_lower or "depot" in name_lower:
            return "depot"

    # Fallback : détection par nom uniquement
    return detect_pdf_type(filename)


def _build_metadata_excel_computed(computed: ComputedMetadata) -> bytes:
    """Génère l'Excel des métadonnées en mémoire avec les données métier."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="Métadonnées")
    meta_rows = computed_metadata_to_rows(computed)
    write_metadata_sheet(ws, meta_rows)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _build_annexe_excel(datasets: list[Dataset]) -> bytes:
    """Génère l'Excel des tableaux en mémoire."""
    wb = Workbook()
    wb.remove(wb.active)
    for ds in datasets:
        sheet_name = _sanitize_sheet_name(ds.name)
        ws = wb.create_sheet(title=sheet_name)
        write_dataset_to_sheet(ws, ds)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def extract_zip_to_dossiers(zip_bytes: bytes | io.BytesIO) -> list[DemandeDossier]:
    """Extrait un ZIP et regroupe les PDF par préfixe numérique.

    Gère la structure mails/proof/ ou les PDFs à plat.
    """
    if isinstance(zip_bytes, bytes):
        zip_bytes = io.BytesIO(zip_bytes)

    dossiers: dict[str, DemandeDossier] = {}

    with zipfile.ZipFile(zip_bytes, "r") as zf:
        for entry in zf.namelist():
            # Ignorer les dossiers et fichiers non-PDF
            if entry.endswith("/") or not entry.lower().endswith(".pdf"):
                continue

            # Ignorer les fichiers __MACOSX
            if "__MACOSX" in entry:
                continue

            filename = Path(entry).name
            parent_dir = str(Path(entry).parent)
            prefix = _extract_prefix(filename)

            if not prefix:
                logger.warning("Pas de prefixe numerique pour %s, ignore", filename)
                continue

            pdf_data = zf.read(entry)
            file_type = _classify_file(filename, parent_dir)

            if prefix not in dossiers:
                dossiers[prefix] = DemandeDossier(prefix=prefix)

            d = dossiers[prefix]
            if file_type == "courrier":
                d.courrier_pdf = pdf_data
                d.courrier_filename = filename
            elif file_type == "ar":
                d.ar_pdf = pdf_data
                d.ar_filename = filename
            elif file_type == "depot":
                d.depot_pdf = pdf_data
                d.depot_filename = filename
            else:
                # Fichier inconnu — essayer comme courrier s'il n'y en a pas déjà un
                if d.courrier_pdf is None:
                    d.courrier_pdf = pdf_data
                    d.courrier_filename = filename
                logger.info("Type inconnu pour %s (prefix %s)", filename, prefix)

    result = sorted(dossiers.values(), key=lambda d: d.prefix)
    logger.info("ZIP extrait : %d demande(s) detectee(s)", len(result))
    return result


def process_demande(
    dossier: DemandeDossier, min_cols: int = MIN_COLS_DEFAULT
) -> DemandeResult:
    """Traite une demande complète : métadonnées (2 étapes) + tableaux."""
    result = DemandeResult(prefix=dossier.prefix)

    # Collecter les PDFs sources
    if dossier.courrier_pdf:
        result.source_pdfs["courrier"] = dossier.courrier_pdf
        result.source_filenames["courrier"] = dossier.courrier_filename
    if dossier.ar_pdf:
        result.source_pdfs["ar"] = dossier.ar_pdf
        result.source_filenames["ar"] = dossier.ar_filename
    if dossier.depot_pdf:
        result.source_pdfs["depot"] = dossier.depot_pdf
        result.source_filenames["depot"] = dossier.depot_filename

    # --- Etape 1 : Extraction brute ---
    try:
        raw = build_raw_metadata(
            courrier_bytes=dossier.courrier_pdf,
            ar_bytes=dossier.ar_pdf,
            depot_bytes=dossier.depot_pdf,
            courrier_filename=dossier.courrier_filename or "",
        )
        result.raw_metadata = raw
    except Exception as e:
        logger.error("Erreur extraction brute prefix %s: %s", dossier.prefix, e)
        raw = RawMetadata()

    # Fallback N° demande
    result.numero_demande = raw.numero_demande or dossier.prefix

    # Remplir aussi le DossierMetadata legacy pour compatibilite
    metadata = DossierMetadata(
        numero_demande=result.numero_demande,
        objet=raw.objet_complet,
        libelle=raw.adresses,
        motif_vacance=raw.motif_vacance,
        date_courrier=raw.date_courrier,
        responsable=raw.responsable,
        numero_lr_depot=raw.numero_lr_depot,
        numero_lr_ar=raw.numero_lr_ar,
        date_presentation_ar=raw.date_presentation_ar,
        date_distribution_ar=raw.date_distribution_ar,
    )
    if dossier.courrier_filename:
        metadata.type_fichiers["courrier"] = dossier.courrier_filename
    if dossier.ar_filename:
        metadata.type_fichiers["ar"] = dossier.ar_filename
    if dossier.depot_filename:
        metadata.type_fichiers["depot"] = dossier.depot_filename
    result.metadata = metadata

    # --- Extraction des tableaux du courrier ---
    if dossier.courrier_pdf:
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        try:
            tmp.write(dossier.courrier_pdf)
            tmp.close()
            tmp_path = Path(tmp.name)

            scan_result = scan_pdf(tmp_path, min_cols=min_cols)
            if scan_result.tables:
                result.table_count = len(scan_result.tables)
                datasets = process_tables(scan_result.tables)
                result.datasets = datasets
                result.row_count = sum(
                    len(ds.data_rows) + len(ds.total_rows) for ds in datasets
                )
                if datasets:
                    result.annexe_excel = _build_annexe_excel(datasets)

        except Exception as e:
            logger.error("Erreur extraction tableaux prefix %s: %s", dossier.prefix, e)
            result.error = str(e)
        finally:
            Path(tmp.name).unlink(missing_ok=True)

    # --- Extraction des donnees des colonnes du tableau ---
    table_data = None
    if result.datasets:
        try:
            table_data = extract_table_data(result.datasets)
        except Exception as e:
            logger.error("Erreur extraction colonnes prefix %s: %s", dossier.prefix, e)

    # --- Etape 2 : Transformation en donnees metier ---
    try:
        computed = compute_metadata(
            raw,
            datasets=result.datasets,
            courrier_bytes=dossier.courrier_pdf,
            table_data=table_data,
            prefix=dossier.prefix,
        )
        result.computed_metadata = computed
    except Exception as e:
        logger.error("Erreur transformation metadata prefix %s: %s", dossier.prefix, e)
        computed = ComputedMetadata()

    # Generer l'Excel metadonnees avec les donnees metier
    try:
        result.metadata_excel = _build_metadata_excel_computed(computed)
    except Exception as e:
        logger.error("Erreur generation metadonnees Excel prefix %s: %s", dossier.prefix, e)

    logger.info(
        "Demande %s : N°%s, %d tableaux, %d lignes",
        dossier.prefix, result.numero_demande, result.table_count, result.row_count,
    )

    return result


def process_zip(
    zip_bytes: bytes | io.BytesIO,
    min_cols: int = MIN_COLS_DEFAULT,
    on_progress: callable = None,
) -> list[DemandeResult]:
    """Pipeline complet : ZIP → liste de résultats par demande.

    Args:
        zip_bytes: Contenu du fichier ZIP.
        min_cols: Seuil minimum de colonnes.
        on_progress: Callback(current, total, prefix) pour la progression.

    Returns:
        Liste de DemandeResult.
    """
    dossiers = extract_zip_to_dossiers(zip_bytes)

    if not dossiers:
        logger.warning("Aucune demande trouvee dans le ZIP.")
        return []

    results = []
    for idx, dossier in enumerate(dossiers):
        if on_progress:
            on_progress(idx, len(dossiers), dossier.prefix)

        try:
            result = process_demande(dossier, min_cols=min_cols)
            results.append(result)
        except Exception as e:
            logger.error("Erreur traitement demande %s: %s", dossier.prefix, e)
            results.append(DemandeResult(
                prefix=dossier.prefix,
                error=str(e),
            ))

    if on_progress:
        on_progress(len(dossiers), len(dossiers), "")

    return results


def _build_recapitulatif_excel(results: list[DemandeResult]) -> bytes:
    """Génère un Excel récapitulatif avec une ligne par demande."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(title="Recapitulatif")

    # Collecter les rows de chaque demande
    all_rows = []
    for r in results:
        meta_rows = computed_metadata_to_rows(r.computed_metadata)
        all_rows.append(meta_rows)

    if not all_rows:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.getvalue()

    # Headers (ligne 1)
    headers = [key for key, _ in all_rows[0]]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Données (une ligne par demande)
    for row_num, meta_rows in enumerate(all_rows, start=2):
        for col_idx, (key, value) in enumerate(meta_rows, start=1):
            if value is None or value == "":
                cell = ws.cell(row=row_num, column=col_idx, value=None)
            else:
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                if isinstance(value, float):
                    cell.number_format = EURO_FORMAT
            cell.font = META_VAL_FONT
            cell.alignment = CELL_ALIGNMENT
            cell.border = THIN_BORDER

    # Auto-ajustement largeur
    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(header))
        for meta_rows in all_rows[:50]:
            _, val = meta_rows[col_idx - 1]
            if val is not None and val != "":
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

    ws.freeze_panes = "A2"

    # Auto-filtre
    if headers:
        last_col = get_column_letter(len(headers))
        last_row = 1 + len(all_rows)
        ws.auto_filter.ref = f"A1:{last_col}{last_row}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_output_zip(results: list[DemandeResult]) -> bytes:
    """Construit le ZIP de sortie structuré par demande.

    Structure :
        Demande_XXX/
            XXX-Courrier_xxx.pdf
            XXX-AR_xxx.pdf
            XXX-Preuve_xxx.pdf
            XXX_Annexe_Tableaux.xlsx
            XXX_Métadonnées.xlsx
    """
    buf = io.BytesIO()

    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for result in results:
            folder = f"Demande_{result.prefix}"

            # Copier les PDF originaux
            for file_type, pdf_data in result.source_pdfs.items():
                filename = result.source_filenames.get(file_type, f"{result.prefix}_{file_type}.pdf")
                zf.writestr(f"{folder}/{filename}", pdf_data)

            # Excel des tableaux
            if result.annexe_excel:
                zf.writestr(
                    f"{folder}/{result.prefix}_Annexe_Tableaux.xlsx",
                    result.annexe_excel,
                )

            # Excel des métadonnées
            if result.metadata_excel:
                zf.writestr(
                    f"{folder}/{result.prefix}_Métadonnées.xlsx",
                    result.metadata_excel,
                )

        # Récapitulatif global à la racine du ZIP
        if len(results) > 0:
            recap_bytes = _build_recapitulatif_excel(results)
            zf.writestr("Recapitulatif_Metadonnees.xlsx", recap_bytes)

    buf.seek(0)
    return buf.getvalue()
