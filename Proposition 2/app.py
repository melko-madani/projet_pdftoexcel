import io
import re
import pdfplumber
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# --- Configuration ---
MOTS_CLES_OBJET = [
    "Dégrèvement pour Travaux",
    "Autre Régularisation",
    "Vacance",
    "Régularisation Abattement/Exonération",
    "Accessibilité PMR",
    "Travaux",
    "ElementsDeConfort",
    "Economie d'énergie",
    "Démolition",
    "Coefficients",
    "Autre",
    "Locative",
    "TypeDeBien",
    "HorsPatrimoine",
    "FinDeGestion",
    "VideOrdures",
    "Categorie",
    "Adresse",
    "ThLogementsVacants",
]


def clean_text(text):
    if not text:
        return ""
    return " ".join(str(text).split())


def detect_mots_cles(page_text):
    """Détecte la présence des mots-clés dans le texte de la première page."""
    resultats = {}
    text_lower = page_text.lower() if page_text else ""
    for mot in MOTS_CLES_OBJET:
        resultats[mot] = "Oui" if mot.lower() in text_lower else "Non"
    return resultats


def extract_metadata(first_page_text):
    """Extrait Libellé, Type et Catégorie de la première page."""
    metadata = {
        "Libellé de la Demande": "N/A",
        "Type": "N/A",
        "Catégorie": "N/A",
    }
    objet_match = re.search(
        r"Objet\s*:\s*Demande de\s+(.*?)\s+pour vacance",
        first_page_text or "",
        re.IGNORECASE | re.DOTALL,
    )
    if objet_match:
        metadata["Libellé de la Demande"] = clean_text(objet_match.group(1))
        if not metadata["Libellé de la Demande"].lower().startswith("dégrèvement"):
            metadata["Libellé de la Demande"] = (
                "dégrèvement " + metadata["Libellé de la Demande"]
            )
    if first_page_text and "vacance" in first_page_text.lower():
        metadata["Type"] = "vacance"
    if first_page_text and (
        "locatif" in first_page_text.lower() or "locative" in first_page_text.lower()
    ):
        metadata["Catégorie"] = "locative"
    return metadata


def process_pdf(pdf_file):
    """Traite un fichier PDF et retourne les données extraites."""
    with pdfplumber.open(pdf_file) as pdf:
        # Page 1 : texte pour mots-clés et métadonnées
        first_page_text = pdf.pages[0].extract_text() if pdf.pages else ""
        mots_cles = detect_mots_cles(first_page_text)
        metadata = extract_metadata(first_page_text)

        # Toutes les pages : extraction des tableaux
        table_rows = []
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if table and len(table) > 1:
                    headers = [clean_text(h) for h in table[0]]
                    for row in table[1:]:
                        clean_row = [clean_text(cell) for cell in row]
                        row_dict = {}
                        row_dict["Libellé de la Demande"] = metadata[
                            "Libellé de la Demande"
                        ]
                        row_dict["Type"] = metadata["Type"]
                        row_dict["Catégorie"] = metadata["Catégorie"]
                        for i, val in enumerate(clean_row):
                            col_name = headers[i] if i < len(headers) else f"Col_{i+1}"
                            row_dict[col_name] = val
                        table_rows.append(row_dict)

    return mots_cles, metadata, table_rows


def style_excel(wb):
    """Applique un style professionnel au workbook."""
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for ws in wb.worksheets:
        # Style des en-têtes
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # Style des données + auto-width
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max(max_length + 2, 10), 40)


def generate_excel(objet_data, tables_data):
    """Génère le fichier Excel avec les 2 sheets."""
    wb = Workbook()

    # --- Sheet 1 : Objet de Courrier ---
    ws_objet = wb.active
    ws_objet.title = "Objet de Courrier"

    if objet_data:
        headers = ["Fichier"] + MOTS_CLES_OBJET
        ws_objet.append(headers)
        for row in objet_data:
            ws_objet.append(row)

    # --- Sheet 2 : Tableaux ---
    ws_tables = wb.create_sheet("Tableaux")

    if tables_data:
        df_tables = pd.DataFrame(tables_data)
        # En-têtes
        ws_tables.append(list(df_tables.columns))
        # Données
        for _, row in df_tables.iterrows():
            ws_tables.append(list(row.values))

    style_excel(wb)

    # Sauvegarder en mémoire
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# =====================
# Interface Streamlit
# =====================
st.set_page_config(
    page_title="Extraction Courrier PDF",
    page_icon="📄",
    layout="wide",
)

st.title("📄 Extraction de données PDF → Excel")
st.markdown("Uploadez vos courriers PDF pour extraire les données dans un fichier Excel.")

uploaded_files = st.file_uploader(
    "Choisir un ou plusieurs fichiers PDF",
    type=["pdf"],
    accept_multiple_files=True,
)

if uploaded_files:
    st.info(f"{len(uploaded_files)} fichier(s) sélectionné(s)")

    if st.button("🚀 Extraire les données", type="primary"):
        all_objet_rows = []
        all_table_rows = []

        progress = st.progress(0)
        status = st.empty()

        for i, pdf_file in enumerate(uploaded_files):
            status.text(f"Traitement de : {pdf_file.name}...")
            try:
                mots_cles, metadata, table_rows = process_pdf(pdf_file)

                # Ligne pour la sheet Objet de Courrier
                objet_row = [pdf_file.name] + [
                    mots_cles[mot] for mot in MOTS_CLES_OBJET
                ]
                all_objet_rows.append(objet_row)

                # Lignes pour la sheet Tableaux
                for row in table_rows:
                    row["Fichier Source"] = pdf_file.name
                all_table_rows.extend(table_rows)

            except Exception as e:
                st.error(f"Erreur sur {pdf_file.name} : {e}")

            progress.progress((i + 1) / len(uploaded_files))

        status.text("Génération du fichier Excel...")

        # Générer Excel
        excel_buffer = generate_excel(all_objet_rows, all_table_rows)

        status.success(
            f"Extraction terminée ! {len(all_objet_rows)} PDF traité(s), "
            f"{len(all_table_rows)} lignes de tableaux extraites."
        )

        # Aperçu des résultats
        st.subheader("Aperçu — Objet de Courrier")
        if all_objet_rows:
            df_objet = pd.DataFrame(
                all_objet_rows, columns=["Fichier"] + MOTS_CLES_OBJET
            )
            st.dataframe(df_objet, use_container_width=True)
        else:
            st.warning("Aucune donnée extraite pour l'objet de courrier.")

        st.subheader("Aperçu — Tableaux extraits")
        if all_table_rows:
            df_tables = pd.DataFrame(all_table_rows)
            st.dataframe(df_tables, use_container_width=True)
        else:
            st.warning("Aucun tableau extrait.")

        # Bouton de téléchargement
        st.download_button(
            label="📥 Télécharger le fichier Excel",
            data=excel_buffer,
            file_name="extraction_courrier.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="primary",
        )
